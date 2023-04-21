import re
import time
import xlsxwriter
import os
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook, load_workbook
import requests
from bs4 import BeautifulSoup

def gather_and_write(link, filename, city):
    fields = ["Position", "Rate", "Cantidate Qualifications", "City"]
    if os.path.exists(str(filename) + ".xlsx"):
        wb = load_workbook(str(filename)+'.xlsx')
        worksheet = wb.active
        rowNum = worksheet.max_row + 1
    else:
        wb = Workbook()
        worksheet = wb.active
        worksheet.title = "Sheet"
        for i in range(0, 4):
            worksheet.cell(row=1, column=i+1, value=str(fields[i]))
            rowNum = 2

    #send request to web page
    response = requests.get(link)

    #parse html content
    soup = BeautifulSoup(response.content, "html.parser")
    
    regex = r"\$[\d\.,]+[\+\-]*\s*(?:to|-)\s*\$[\d\.,]+[\+\-]*|\$[\d\.,]+[\+\-]*"
    # Search for the regex in the HTML content and extract the first match
    match = re.search(regex, soup.decode())
    salary = "[Not Found]"
    # Extract the matched string from the match object
    if match:
        salary = match.group()
        

    position_element = soup.title.text
    qualy = []
    choice = False
# find the qualifications
    elements = soup.find_all(class_="ats-description")

    for item in elements:
        for i in str(item.text).split(" "):
            try:
                if i[0:15] == "qualifications:":
                    choice = True
                    qualy.append(i)
                    continue
                else:
                    if choice:
                        qualy.append(i)
                    continue
            except:
                continue
    qualy = ' '.join(qualy)
    list = [position_element, salary, qualy, city]
    
    #write to excel 

    #print data list across row
    for i in range(0 ,4):
        try:
            worksheet.cell(row=rowNum, column=i+1, value=str(list[i]))
        except:
            break
    # save the workbook
    wb.save(name)

filename = input("Enter a filename: ")
name = str(filename) + ".xlsx"
options = Options()
options.headless = False
driver = webdriver.Chrome(options=options)
realUrl = "https://jobs.merakey.org/search-jobs/"
driver.get(realUrl)
totcounter = 0
link_counter = 1
#FIND HOW MANY
element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "/html/body/div[2]/main/div[1]/section/h1")))
text = element.text
numbers = re.findall(r'\d+', text)
titles = int(numbers[0])
while True:
    #get link
    print(f"Titles collected: {totcounter} out of {titles}.", end='\r')
    try:
        link = WebDriverWait(driver, 5).until(lambda driver: driver.find_element(By.XPATH, "/html/body/div[2]/main/div[1]/section/section/ul/li["+str(link_counter)+"]/a").get_attribute('href'))
        city = WebDriverWait(driver, 5).until(lambda driver: driver.find_element(By.XPATH, "/html/body/div[2]/main/div[1]/section/section/ul/li["+str(link_counter)+"]/a/span[1]")).text                                                                                                                                 
    except:
    #give time so not get ban click next if no more link
        try:
            time.sleep(1.22)
            link_counter = 1
            driver.execute_script("arguments[0].click();", WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/main/div[1]/section/section/nav/div[2]/a[2]"))))
            link = driver.find_element(by=By.XPATH, value="/html/body/div[2]/main/div[1]/section/section/ul/li["+str(link_counter)+"]/a").get_attribute("href")

            #forcing error when hit limit otherwise can keep going
            if titles < totcounter:
                forcingerrors = 1/0
        
        except:
            print("completed")
            break

    #time for inconsistent data output?
    time.sleep(.1)
    #send to write
    gather_and_write(link, filename, city)
     #time for inconsistent data output?
    time.sleep(.1)
    #endings counting
    link_counter = link_counter + 1
    totcounter = totcounter + 1

print("complete closing console")
time.sleep(10)
quit()
